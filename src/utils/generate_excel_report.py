import json
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


# Function to load JSON from a file
def load_json(file_path):
    with open(file_path, "r") as file:
        return json.load(file)


# Function to generate Excel report with custom formatting
def generate_excel_report(json_file_path, output_path):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Report"

    data = load_json(json_file_path)

    # Define headers and format
    headers = [
        "Project Name",
        "Issue",
        "Issue Details",
        "Failed Checks Summary",
        "Passed Checks Summary",
    ]
    ws.append(headers)

    # Convert pixel to Excel width: pixels / 7
    column_width_in_pixels = 450
    excel_column_width = column_width_in_pixels / 7

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True, size=12)
        ws.column_dimensions[get_column_letter(col_num)].width = excel_column_width

    # Define styles for formatting
    red_bold_underline = Font(color="FF0000", bold=True, underline="single")
    green_bold_underline = Font(color="008000", bold=True, underline="single")
    default_font = Font(color="000000")  # Default black font for other text
    bullet_prefix = "• "

    # Populate the sheet with project data
    for project in data:
        project_name = (
            f'=HYPERLINK("{project["project_url"]}", "{project["project_name"]}")'
        )
        issue = f'=HYPERLINK("{project["issue_url"]}", "{project["issue"]}")'

        # Format issue details with bullet points
        issue_details = f"{bullet_prefix}" + project["issue_details"].replace(
            ", ", f"\n{bullet_prefix}"
        )

        # Format failed checks summary
        failed_summary = ""
        for check in project["failed_checks"]:
            failed_summary += f"**{check['name']}**\n{bullet_prefix}{check['message']}\n{bullet_prefix}{check['recommendation']}\n\n"

        # Format passed checks summary
        passed_summary = ""
        for check in project["passed_checks"]:
            passed_summary += (
                f"**{check['name']}**\n{bullet_prefix}{check['message']}\n\n"
            )

        # Append data to worksheet
        ws.append([project_name, issue, issue_details, failed_summary, passed_summary])

    # Adjust cell styling and hyperlink formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Apply hyperlink to Project Name and Issue columns
            if cell.col_idx == 1 or cell.col_idx == 2:
                cell.hyperlink = Hyperlink(
                    ref=cell.coordinate, target=cell.value.split('"')[1]
                )

            # Apply formatting for failed and passed check titles only
            if cell.col_idx in [4, 5] and cell.value:
                # Split the cell text into lines
                lines = cell.value.split("\n")
                formatted_text = []

                # Apply formatting to lines that contain titles
                for line in lines:
                    if line.startswith("**"):
                        # Check if the line is a title (starts with '**')
                        if cell.col_idx == 4:
                            cell.font = red_bold_underline
                        else:
                            cell.font = green_bold_underline
                    else:
                        cell.font = default_font
                    formatted_text.append(line)

                # Join the formatted text back into a single string
                cell.value = "\n".join(formatted_text)

    # Set row height in pixels: 409 pixels (each Excel row height unit = 0.75 points)
    row_height_in_pixels = 409
    excel_row_height = row_height_in_pixels / 0.75  # Convert to Excel's points

    # Apply row height and alignment
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = excel_row_height

    # Save the formatted Excel report
    wb.save(output_path)
